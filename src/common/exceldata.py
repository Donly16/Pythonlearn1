
import xlrd

import openpyxl

#import xdrlib ,sys  
def open_excel(file= 'file.xls'):  
        try:  
                data = xlrd.open_workbook(file)  
                #data = openpyxl.load_workbook(file)
                return data  
        except Exception as E:  
                print("can't open file:{}".format(file))   
                raise(E)

#根据工作表的名称或索引获取Excel表格中的数据
#参数:file：Excel文件路径
#     rowindex：要取的数据所在行的索引+1，不指定或为0，则取整个表的值
#     rowname ：行名，即第一列的值。如果有指定行名，则优先按照行名来取数据
#     by_index:工作表的索引，默认为0，即第一个表
#     by_name：Sheet1名称，如果同时指定了工作表名和索引，则优先按照工作表名来取
#返回值：包含所有行或指定行数据字典的列表
def excel_read_row(file= 'file.xls',rowindex=0,rowname = 'null',by_index=0,by_name='null'):
    data = open_excel(file)

    #如果不指定sheet名字，则默认按索引来取表格
    if by_name != 'null':
            table = data.sheet_by_name(by_name)
    else:        
            table = data.sheets()[by_index]    

    
    list =[]
    nrows = table.nrows #行数
    if nrows == 0:
        #raise Exception("there is no data in the sheet")
        return list
    #按行名查找行索引
    if rowname != 'null':
        for i in range(nrows):
            print(table.row_values(i)[0])
            if rowname == table.row_values(i)[0]:
                rowindex = i + 1                
                break
        else:
            return list
    
    colnames =  table.row_values(0) #第一行数据

    #获取要取的数据的行数，如果没有指定，则取表格所有数据
    start = (rowindex - 1) if rowindex > 0 else 1
    end = (start + 1)if rowindex >0 else nrows
       
    for rownum in range(start,end): 
        row = table.row_values(rownum)
        if row:
            app = {}
            for i in range(len(colnames)):                
                app[colnames[i]] = row[i]
        list.append(app)
        
    return list

#根据表索引获取Excel表格中的数据
#参数:file：Excel文件路径,
#     colindex：要取的数据列索引+1,不指定或为0，则取整个表的值
#     by_index:工作表的索引，默认为0，即第一个表
#     by_name：Sheet1名称，如果同时指定了工作表名和索引，则优先按照工作表名来取
#     col_name:列名，如果同时指定了列索引和列名，则优先按照列名来取
#返回值：如果不指定列索引，则返回包含每行数据字典的列表如果指定了列索引，则返回一个包含列数据的列表
def excel_read_col(file= 'file.xls',colindex=0,by_index=0,by_name='null',col_name='null'):  
    data = open_excel(file)

    #如果不指定sheet名字，则默认按索引来取表格
    if by_name != 'null':
            table = data.sheet_by_name(by_name)
    else:        
            table = data.sheets()[by_index]    


    list =[]
        
    nrows = table.nrows #行数
    ncols = table.ncols #列数
    if nrows == 0:
        #raise Exception("there is no data in the sheet")
        return list
        
    colnames = table.row_values(0) #第一行标题

    #查找列名指定列的索引，并加1
    if col_name != 'null':
        colindex = colnames.index(col_name)+1

        
    #获取要取的数据的列，如果没有指定，则取表格所有数据
    start = (colindex - 1)if colindex >0 else 0
    end = (colindex or ncols)

    #遍历所有行
    for rownum in range(1,nrows):  
        row = table.row_values(rownum)

        if row:  
            app = {}
            #遍历所有列或者指定列
            for i in range(start,end):
                if colindex > 0:
                    app = row[i]
                    #list.append(row[i])
                else:
                    app[colnames[i]] = row[i]
                       
            list.append(app)

    return list 


#根据表索引获取Excel表格中的数据
#参数:file：Excel文件路径,
#     col：要取的数据列索引+1
#     row:要取的数据行索引+1
#     by_index:工作表的索引，默认为0，即第一个表
#     by_name：Sheet1名称，如果同时指定了工作表名和索引，则优先按照工作表名来取
#返回值：单元格的值,不指定则默认返回第一行或第一列的值
def excel_read_cell(file= 'file.xls',row=0,col=0,by_index=0,by_name='null'):  
    data = open_excel(file)

    #如果不指定sheet名字，则默认按索引来取表格
    if by_name != 'null':
            table = data.sheet_by_name(by_name)
    else:        
            table = data.sheets()[by_index]    
    row = (row - 1)if row > 0 else 0
    col = (col - 1)if col > 0 else 0
    return table.cell(row,col).value


#向已存在的excel末尾追加一行
#参数:file：Excel文件路径,
#     colnames:列名，列表形式,不能为空，如果不指定，则默认有5列
#     by_name：默认Sheet1名称
#返回值：
def new_excel(file= 'file.xls',colnames=['col1','col2','col3','col4','col5'],by_name=u'Sheet1'):  
    wb = openpyxl.Workbook()
    table = wb.create_sheet(title=by_name,index=0)   

    table.append(colnames)
        
    wb.save(file)

    return


#根据表索引获取Excel表格中的数据
#参数:file：Excel文件路径,
#     col：要取的数据列索引
#     row:要取的数据行索引
#     value:要填入单元格的值
#     by_name：默认Sheet1名称
#返回值：
def excel_write_cell(file= 'file.xls',row=0,col=0,value='',by_name=u'Sheet1'):  
    data = openpyxl.load_workbook(file)
    
    table =data.get_sheet_by_name(by_name)   

    table.cell(row = row,column= col).value=value
    
    data.save(file)

    return 

#向已存在的excel末尾追加一行
#参数:file：Excel文件路径,
#     value:要填入行的值，列表形式
#     by_name：默认Sheet1名称
#返回值：
def excel_append_row(file= 'file.xls',value=[],by_name=u'Sheet1'):  
    data = openpyxl.load_workbook(file)
    
    table =data.get_sheet_by_name(by_name)   

    table.append(value)
    
    data.save(file)

    return 

#if __name__ == '__main__':  
#    test()  
